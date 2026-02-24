"""Tests for output generators (Excel, HTML, PDF, JSON)."""
import json
import pytest
from decimal import Decimal
from pathlib import Path

from src.models import AnalysisResult, MappedData
from src.calc import (
    calculate_income_statement,
    calculate_working_capital,
    estimate_balance_sheet,
    calculate_cash_flow,
    calculate_ratios,
    calculate_power_of_one,
    classify_cash_quality,
)
from src.output.excel_report import ExcelReportGenerator
from src.output.html_dashboard import HTMLDashboardGenerator
from src.output.json_export import JSONExporter


# ---------------------------------------------------------------------------
# Helpers shared across test classes
# ---------------------------------------------------------------------------

def _build_mapped() -> MappedData:
    """Build a minimal MappedData for testing, mirroring the Q1 conftest data."""
    return MappedData(
        company="Test Company",
        period="Q1_2025",
        period_type="quarter",
        days_in_period=90,
        gross_revenue=Decimal("40100000.00"),
        returns_deductions=Decimal("2500000.00"),
        cogs=Decimal("30650000.00"),
        operating_expenses=Decimal("19650000.00"),
        financial_expenses=Decimal("1200000.00"),
        financial_income=Decimal("150000.00"),
        accounts_receivable=Decimal("18500000.00"),
        inventory=Decimal("3200000.00"),
        accounts_payable=Decimal("8900000.00"),
        cash=Decimal("1200000.00"),
        short_term_debt=Decimal("12000000.00"),
        long_term_debt=Decimal("25000000.00"),
        shareholders_equity=Decimal("35000000.00"),
        ppe_gross=Decimal("45000000.00"),
    )


def _run_calc_chain(mapped: MappedData):
    """Run the full calculation chain on a MappedData, returning a PeriodResult."""
    is_result = calculate_income_statement(mapped)
    wc_result = calculate_working_capital(mapped, days_in_period=mapped.days_in_period)
    merged = is_result.model_copy(update={
        "days_sales_outstanding": wc_result.days_sales_outstanding,
        "days_inventory_outstanding": wc_result.days_inventory_outstanding,
        "days_payable_outstanding": wc_result.days_payable_outstanding,
        "cash_conversion_cycle": wc_result.cash_conversion_cycle,
        "working_capital": wc_result.working_capital,
        "working_capital_investment": wc_result.working_capital_investment,
    })
    bs_result = estimate_balance_sheet(merged)
    cf_result = calculate_cash_flow(bs_result)
    return calculate_ratios(cf_result)


def _make_analysis_result() -> AnalysisResult:
    """Build a minimal AnalysisResult for testing via the full calc chain."""
    mapped = _build_mapped()
    pr = _run_calc_chain(mapped)
    levers = calculate_power_of_one(pr)
    cq = classify_cash_quality(pr)
    return AnalysisResult(
        company="Test Company",
        periods=[pr],
        power_of_one=levers,
        cash_quality=cq,
    )


# ---------------------------------------------------------------------------
# ExcelReportGenerator
# ---------------------------------------------------------------------------

class TestExcelReportGenerator:
    """Tests for ExcelReportGenerator."""

    def test_generates_xlsx_file(self, tmp_path):
        """Generator creates a non-empty .xlsx file at the given path."""
        path = tmp_path / "report.xlsx"
        result = _make_analysis_result()
        ExcelReportGenerator(str(path)).generate(result)
        assert path.exists()
        assert path.stat().st_size > 0

    def test_returns_path_object(self, tmp_path):
        """generate() returns the Path of the created file."""
        path = tmp_path / "report.xlsx"
        result = _make_analysis_result()
        returned = ExcelReportGenerator(str(path)).generate(result)
        assert Path(returned) == path

    def test_creates_parent_directories(self, tmp_path):
        """Generator creates missing parent directories automatically."""
        path = tmp_path / "sub" / "dir" / "report.xlsx"
        result = _make_analysis_result()
        ExcelReportGenerator(str(path)).generate(result)
        assert path.exists()

    def test_generates_valid_xlsx(self, tmp_path):
        """Generated file is a valid xlsx workbook (openpyxl can open it)."""
        pytest.importorskip("openpyxl")
        import openpyxl

        path = tmp_path / "report.xlsx"
        result = _make_analysis_result()
        ExcelReportGenerator(str(path)).generate(result)
        wb = openpyxl.load_workbook(str(path))
        assert len(wb.sheetnames) > 0

    def test_workbook_contains_expected_sheets(self, tmp_path):
        """Generated workbook has the canonical sheet names."""
        pytest.importorskip("openpyxl")
        import openpyxl

        expected_sheets = {
            "Resumo Executivo",
            "Cap 1 Rentabilidade",
            "Cap 2 Capital de Giro",
            "Cap 3 Outros Capitais",
            "Cap 4 Financiamento",
            "Power of One",
            "DRE",
            "Fluxo de Caixa",
            "Indicadores",
            "Qualidade do Caixa",
        }
        path = tmp_path / "report.xlsx"
        result = _make_analysis_result()
        ExcelReportGenerator(str(path)).generate(result)
        wb = openpyxl.load_workbook(str(path))
        assert expected_sheets.issubset(set(wb.sheetnames))

    def test_handles_multiple_periods(self, tmp_path):
        """Generator works when AnalysisResult has more than one period."""
        pytest.importorskip("openpyxl")
        mapped = _build_mapped()
        pr1 = _run_calc_chain(mapped)
        pr2 = _run_calc_chain(mapped)
        result = AnalysisResult(
            company="Test Company",
            periods=[pr1, pr2],
            power_of_one=calculate_power_of_one(pr1),
            cash_quality=classify_cash_quality(pr1),
        )
        path = tmp_path / "multi.xlsx"
        ExcelReportGenerator(str(path)).generate(result)
        assert path.exists() and path.stat().st_size > 0

    def test_raises_import_error_without_openpyxl(self, tmp_path, monkeypatch):
        """generate() raises ImportError when openpyxl is not installed."""
        import src.output.excel_report as excel_mod

        monkeypatch.setattr(excel_mod, "_OPENPYXL_AVAILABLE", False)
        path = tmp_path / "report.xlsx"
        result = _make_analysis_result()
        with pytest.raises(ImportError, match="openpyxl"):
            ExcelReportGenerator(str(path)).generate(result)


# ---------------------------------------------------------------------------
# HTMLDashboardGenerator
# ---------------------------------------------------------------------------

class TestHTMLDashboardGenerator:
    """Tests for HTMLDashboardGenerator."""

    def test_generates_html_file(self, tmp_path):
        """Generator creates a non-empty .html file at the given path."""
        path = tmp_path / "dashboard.html"
        result = _make_analysis_result()
        HTMLDashboardGenerator(str(path)).generate(result)
        assert path.exists()
        assert path.stat().st_size > 0

    def test_returns_path_object(self, tmp_path):
        """generate() returns the Path of the created file."""
        path = tmp_path / "dashboard.html"
        result = _make_analysis_result()
        returned = HTMLDashboardGenerator(str(path)).generate(result)
        assert Path(returned) == path

    def test_html_contains_company_name(self, tmp_path):
        """Generated HTML embeds the company name from AnalysisResult."""
        path = tmp_path / "dashboard.html"
        result = _make_analysis_result()
        HTMLDashboardGenerator(str(path)).generate(result)
        content = path.read_text(encoding="utf-8")
        assert "Test Company" in content

    def test_html_is_valid_document(self, tmp_path):
        """Generated output starts with an HTML doctype declaration."""
        path = tmp_path / "dashboard.html"
        result = _make_analysis_result()
        HTMLDashboardGenerator(str(path)).generate(result)
        content = path.read_text(encoding="utf-8")
        assert content.strip().lower().startswith("<!doctype html")

    def test_html_contains_period_label(self, tmp_path):
        """Generated HTML includes the period identifier from PeriodResult."""
        path = tmp_path / "dashboard.html"
        result = _make_analysis_result()
        HTMLDashboardGenerator(str(path)).generate(result)
        content = path.read_text(encoding="utf-8")
        assert "Q1_2025" in content

    def test_html_contains_chartjs_script(self, tmp_path):
        """Generated HTML references Chart.js for interactive charts."""
        path = tmp_path / "dashboard.html"
        result = _make_analysis_result()
        HTMLDashboardGenerator(str(path)).generate(result)
        content = path.read_text(encoding="utf-8")
        assert "chart.js" in content.lower()

    def test_creates_parent_directories(self, tmp_path):
        """Generator creates missing parent directories automatically."""
        path = tmp_path / "nested" / "output" / "dashboard.html"
        result = _make_analysis_result()
        HTMLDashboardGenerator(str(path)).generate(result)
        assert path.exists()

    def test_handles_no_ai_insights(self, tmp_path):
        """Generator does not crash when ai_insights is None."""
        path = tmp_path / "dashboard.html"
        mapped = _build_mapped()
        pr = _run_calc_chain(mapped)
        result = AnalysisResult(
            company="Test Company",
            periods=[pr],
        )
        HTMLDashboardGenerator(str(path)).generate(result)
        assert path.exists() and path.stat().st_size > 0

    def test_handles_multiple_periods(self, tmp_path):
        """Generator renders correctly with multiple periods (trend chart enabled)."""
        mapped = _build_mapped()
        pr1 = _run_calc_chain(mapped)
        pr2 = _run_calc_chain(mapped)
        result = AnalysisResult(
            company="Test Company",
            periods=[pr1, pr2],
            power_of_one=calculate_power_of_one(pr1),
            cash_quality=classify_cash_quality(pr1),
        )
        path = tmp_path / "multi.html"
        HTMLDashboardGenerator(str(path)).generate(result)
        content = path.read_text(encoding="utf-8")
        # Multiple periods enables the revenue trend chart canvas
        assert "revenueChart" in content

    def test_optional_template_dir_ignored(self, tmp_path):
        """Passing a template_dir does not break generation (kept for API compat)."""
        path = tmp_path / "dashboard.html"
        result = _make_analysis_result()
        HTMLDashboardGenerator(str(path), template_dir="/nonexistent").generate(result)
        assert path.exists()


# ---------------------------------------------------------------------------
# JSONExporter
# ---------------------------------------------------------------------------

class TestJSONExporter:
    """Tests for JSONExporter."""

    def test_generates_json_file(self, tmp_path):
        """Exporter creates a non-empty .json file at the given path."""
        path = tmp_path / "export.json"
        result = _make_analysis_result()
        JSONExporter(str(path)).export(result)
        assert path.exists()
        assert path.stat().st_size > 0

    def test_returns_path_object(self, tmp_path):
        """export() returns the Path of the created file."""
        path = tmp_path / "export.json"
        result = _make_analysis_result()
        returned = JSONExporter(str(path)).export(result)
        assert Path(returned) == path

    def test_json_is_valid(self, tmp_path):
        """Exported file contains valid, parseable JSON."""
        path = tmp_path / "export.json"
        result = _make_analysis_result()
        JSONExporter(str(path)).export(result)
        data = json.loads(path.read_text(encoding="utf-8"))
        assert isinstance(data, dict)

    def test_json_company_field(self, tmp_path):
        """Exported JSON preserves the company name."""
        path = tmp_path / "export.json"
        result = _make_analysis_result()
        JSONExporter(str(path)).export(result)
        data = json.loads(path.read_text(encoding="utf-8"))
        assert data["company"] == "Test Company"

    def test_json_periods_count(self, tmp_path):
        """Exported JSON contains the correct number of periods."""
        path = tmp_path / "export.json"
        result = _make_analysis_result()
        JSONExporter(str(path)).export(result)
        data = json.loads(path.read_text(encoding="utf-8"))
        assert len(data["periods"]) == 1

    def test_json_decimal_serialization(self, tmp_path):
        """Decimal values are serialized to JSON as numbers or numeric strings
        (Pydantic v2 model_dump with mode='json' may emit strings for Decimal).
        The key requirement is that the value is JSON-parseable as a number."""
        path = tmp_path / "export.json"
        result = _make_analysis_result()
        JSONExporter(str(path)).export(result)
        data = json.loads(path.read_text(encoding="utf-8"))
        net_revenue = data["periods"][0]["net_revenue"]
        # Accept either native JSON number or a numeric string (Pydantic v2 behavior)
        assert isinstance(net_revenue, (int, float, str))
        float(net_revenue)  # Must be convertible to float without error

    def test_json_pretty_printed_by_default(self, tmp_path):
        """Default export is pretty-printed (contains newlines)."""
        path = tmp_path / "export.json"
        result = _make_analysis_result()
        JSONExporter(str(path)).export(result)
        raw = path.read_text(encoding="utf-8")
        assert "\n" in raw

    def test_json_compact_mode(self, tmp_path):
        """pretty=False produces compact JSON without indentation."""
        path = tmp_path / "compact.json"
        result = _make_analysis_result()
        JSONExporter(str(path)).export(result, pretty=False)
        raw = path.read_text(encoding="utf-8")
        # Compact JSON should be a single line (no leading spaces on second line)
        assert not raw.startswith("{\n ")

    def test_json_contains_power_of_one(self, tmp_path):
        """Exported JSON includes the power_of_one levers list."""
        path = tmp_path / "export.json"
        result = _make_analysis_result()
        JSONExporter(str(path)).export(result)
        data = json.loads(path.read_text(encoding="utf-8"))
        assert "power_of_one" in data
        assert isinstance(data["power_of_one"], list)

    def test_json_contains_cash_quality(self, tmp_path):
        """Exported JSON includes the cash_quality metrics list."""
        path = tmp_path / "export.json"
        result = _make_analysis_result()
        JSONExporter(str(path)).export(result)
        data = json.loads(path.read_text(encoding="utf-8"))
        assert "cash_quality" in data
        assert isinstance(data["cash_quality"], list)

    def test_json_period_has_expected_fields(self, tmp_path):
        """Each period object in the JSON contains key financial fields."""
        path = tmp_path / "export.json"
        result = _make_analysis_result()
        JSONExporter(str(path)).export(result)
        data = json.loads(path.read_text(encoding="utf-8"))
        period = data["periods"][0]
        for field in ("net_revenue", "ebitda", "operating_cash_flow", "net_income"):
            assert field in period, f"Expected field '{field}' missing from period JSON"

    def test_creates_parent_directories(self, tmp_path):
        """Exporter creates missing parent directories automatically."""
        path = tmp_path / "nested" / "output" / "export.json"
        result = _make_analysis_result()
        JSONExporter(str(path)).export(result)
        assert path.exists()

    def test_json_generated_at_is_string(self, tmp_path):
        """The generated_at timestamp is serialized as an ISO string."""
        path = tmp_path / "export.json"
        result = _make_analysis_result()
        JSONExporter(str(path)).export(result)
        data = json.loads(path.read_text(encoding="utf-8"))
        assert isinstance(data["generated_at"], str)
        # Basic ISO 8601 sanity check
        assert "T" in data["generated_at"] or "-" in data["generated_at"]
