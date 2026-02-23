"""Excel report generation with multiple sheets."""

from pathlib import Path
from typing import Optional

from src.models import AnalysisResult

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


class ExcelReportGenerator:
    """
    Generates comprehensive Excel report with multiple sheets.

    Sheets included:
    1. Resumo Executivo - Executive summary with Key metrics
    2. Capítulo 1 - Profitability DRE (Income Statement)
    3. Capítulo 2 - Working Capital metrics and trends
    4. Capítulo 3 - Other Capital and CAPEX
    5. Capítulo 4 - Funding and Leverage
    6. Power of One - All 7 levers with impacts
    7. DRE - Detailed Income Statement
    8. Fluxo de Caixa - Cash Flow Statement
    9. Capital de Giro - Working Capital detail
    10. Indicadores - All ratios and metrics

    Formatting:
    - Professional colors (blue headers, light gray alternating rows)
    - Currency formatting (R$ #,##0.00)
    - Percentage formatting (0.00%)
    - Charts for trends and metrics

    Attributes:
        output_path: Path for Excel file
        analysis_result: Complete analysis to export
    """

    # Style constants
    _HEADER_FILL_COLOR = '1F4E79'
    _HEADER_FONT_COLOR = 'FFFFFF'
    _CURRENCY_FORMAT = '#,##0.00'
    _PERCENT_FORMAT = '0.00%'

    def __init__(self, output_path: str) -> None:
        """
        Initialize Excel report generator.

        Args:
            output_path: Path for output Excel file
        """
        self.output_path = Path(output_path)
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self._wb: Optional[object] = None

    def generate(self, analysis_result: AnalysisResult) -> Path:
        """
        Generate complete Excel report.

        Args:
            analysis_result: Complete analysis results to export

        Returns:
            Path: Path to generated Excel file

        Raises:
            ImportError: If openpyxl is not installed
            IOError: If file write fails
        """
        if not _OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required to generate Excel reports. "
                "Install it with: pip install openpyxl"
            )

        self._wb = Workbook()

        self._add_executive_summary_sheet(analysis_result)
        self._add_chapters_sheets(analysis_result)
        self._add_power_of_one_sheet(analysis_result)
        self._add_financial_statements_sheets(analysis_result)

        # Remove the default sheet created by openpyxl
        if "Sheet" in self._wb.sheetnames:
            del self._wb["Sheet"]

        self._wb.save(self.output_path)
        return self.output_path

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _make_header_fill(self) -> "PatternFill":
        return PatternFill(fill_type='solid', fgColor=self._HEADER_FILL_COLOR)

    def _make_header_font(self) -> "Font":
        return Font(bold=True, color=self._HEADER_FONT_COLOR)

    def _write_header_row(self, sheet, row: int, values: list) -> None:
        """Write a styled header row (blue fill, white bold font)."""
        fill = self._make_header_fill()
        font = self._make_header_font()
        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=col_idx, value=value)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _write_label_value_row(
        self,
        sheet,
        row: int,
        label: str,
        values: list,
        number_format: str = _CURRENCY_FORMAT,
    ) -> None:
        """Write a label in column A and period values in subsequent columns."""
        sheet.cell(row=row, column=1, value=label)
        for col_idx, value in enumerate(values, start=2):
            cell = sheet.cell(row=row, column=col_idx, value=float(value) if value is not None else None)
            if value is not None:
                cell.number_format = number_format

    def _format_sheet(self, sheet_name: str) -> None:
        """
        Apply professional formatting to sheet.

        Args:
            sheet_name: Name of sheet to format
        """
        if self._wb is None or sheet_name not in self._wb.sheetnames:
            return

        sheet = self._wb[sheet_name]

        # Set column A (label column) wider, other columns standard
        sheet.column_dimensions['A'].width = 35
        for col_idx in range(2, sheet.max_column + 1):
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = 18

        # Freeze top row
        sheet.freeze_panes = 'A2'

    # ------------------------------------------------------------------
    # Sheet builders
    # ------------------------------------------------------------------

    def _add_executive_summary_sheet(self, analysis_result: AnalysisResult) -> None:
        """
        Add Resumo Executivo (Executive Summary) sheet.

        Content:
        - Company name and period
        - 3 Big Measures (NCF, OCF, MCF) with formatting
        - Overall Cash Quality Grade
        - Top 5 Power of One levers
        - Key metrics summary table
        """
        sheet = self._wb.create_sheet("Resumo Executivo")

        header_fill = self._make_header_fill()
        header_font = self._make_header_font()

        # Row 1: Report title (merged A1:F1)
        title_cell = sheet.cell(row=1, column=1, value="Relatório CashFlow Story")
        title_cell.font = Font(bold=True, size=14, color=self._HEADER_FONT_COLOR)
        title_cell.fill = header_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells('A1:F1')

        # Row 2: Company
        sheet.cell(row=2, column=1, value=f"Empresa: {analysis_result.company}")

        # Row 3: Generated at
        generated_str = analysis_result.generated_at.strftime("%d/%m/%Y %H:%M")
        sheet.cell(row=3, column=1, value=f"Gerado em: {generated_str}")

        # Row 5: Three Big Measures header
        tbm_header = sheet.cell(row=5, column=1, value="3 Grandes Medidas")
        tbm_header.font = header_font
        tbm_header.fill = header_fill

        # Rows 6-8: Three Big Measures values
        tbm = analysis_result.three_big_measures
        measures = [
            ("Fluxo de Caixa Líquido", tbm.net_cash_flow if tbm else None),
            ("Fluxo de Caixa Operacional", tbm.operating_cash_flow if tbm else None),
            ("Fluxo de Caixa Marginal", tbm.marginal_cash_flow if tbm else None),
        ]
        for row_offset, (label, value) in enumerate(measures, start=6):
            sheet.cell(row=row_offset, column=1, value=label)
            val_cell = sheet.cell(
                row=row_offset, column=2,
                value=float(value) if value is not None else None
            )
            if value is not None:
                val_cell.number_format = self._CURRENCY_FORMAT

        # Row 10: Cash Quality header
        cq_header = sheet.cell(row=10, column=1, value="Qualidade do Caixa")
        cq_header.font = header_font
        cq_header.fill = header_fill

        # Cash quality column headers
        self._write_header_row(sheet, row=11, values=["Métrica", "Valor", "Nota"])

        # Cash quality metric rows
        for row_idx, metric in enumerate(analysis_result.cash_quality, start=12):
            sheet.cell(row=row_idx, column=1, value=metric.label_pt)
            val_cell = sheet.cell(row=row_idx, column=2, value=float(metric.value))
            val_cell.number_format = self._CURRENCY_FORMAT
            sheet.cell(row=row_idx, column=3, value=metric.grade)

        self._format_sheet("Resumo Executivo")

    def _add_chapters_sheets(self, analysis_result: AnalysisResult) -> None:
        """
        Add 4 Chapter detail sheets.

        Each sheet contains:
        - Chapter metrics table
        - Period trend columns
        """
        periods = analysis_result.periods
        period_labels = [p.period for p in periods]

        # ---------- Chapter 1: Rentabilidade (Income Statement) ----------
        sheet1 = self._wb.create_sheet("Cap 1 Rentabilidade")
        header_values = ["Métrica"] + period_labels
        self._write_header_row(sheet1, row=1, values=header_values)

        is_rows = [
            ("Receita Bruta", "gross_revenue", self._CURRENCY_FORMAT),
            ("(-) Devoluções e Deduções", "returns_deductions", self._CURRENCY_FORMAT),
            ("Receita Líquida", "net_revenue", self._CURRENCY_FORMAT),
            ("(-) CMV", "cogs", self._CURRENCY_FORMAT),
            ("Lucro Bruto", "gross_profit", self._CURRENCY_FORMAT),
            ("Margem Bruta %", "gross_margin_pct", self._PERCENT_FORMAT),
            ("(-) Despesas Operacionais", "operating_expenses", self._CURRENCY_FORMAT),
            ("EBITDA", "ebitda", self._CURRENCY_FORMAT),
            ("Margem EBITDA %", "ebitda_margin_pct", self._PERCENT_FORMAT),
            ("(-) Depreciação e Amortização", "depreciation_amortization", self._CURRENCY_FORMAT),
            ("EBIT", "ebit", self._CURRENCY_FORMAT),
            ("Margem EBIT %", "ebit_margin_pct", self._PERCENT_FORMAT),
            ("(-) Despesas Financeiras", "financial_expenses", self._CURRENCY_FORMAT),
            ("(+) Receitas Financeiras", "financial_income", self._CURRENCY_FORMAT),
            ("(+/-) Outros", "other_income_expenses", self._CURRENCY_FORMAT),
            ("EBT", "ebt", self._CURRENCY_FORMAT),
            ("(-) IRPJ", "irpj_tax", self._CURRENCY_FORMAT),
            ("(-) CSLL", "csll_tax", self._CURRENCY_FORMAT),
            ("Lucro Líquido", "net_income", self._CURRENCY_FORMAT),
            ("Margem Líquida %", "net_margin_pct", self._PERCENT_FORMAT),
        ]
        for row_idx, (label, field, fmt) in enumerate(is_rows, start=2):
            values = [getattr(p, field) for p in periods]
            self._write_label_value_row(sheet1, row_idx, label, values, fmt)
        self._format_sheet("Cap 1 Rentabilidade")

        # ---------- Chapter 2: Capital de Giro (Working Capital) ----------
        sheet2 = self._wb.create_sheet("Cap 2 Capital de Giro")
        self._write_header_row(sheet2, row=1, values=header_values)

        wc_rows = [
            ("Contas a Receber", "accounts_receivable", self._CURRENCY_FORMAT),
            ("Estoques", "inventory", self._CURRENCY_FORMAT),
            ("Contas a Pagar", "accounts_payable", self._CURRENCY_FORMAT),
            ("DSO (Dias de Recebimento)", "days_sales_outstanding", '#,##0.0'),
            ("DIO (Dias de Estoque)", "days_inventory_outstanding", '#,##0.0'),
            ("DPO (Dias de Pagamento)", "days_payable_outstanding", '#,##0.0'),
            ("Ciclo de Caixa (CCC)", "cash_conversion_cycle", '#,##0.0'),
            ("Capital de Giro", "working_capital", self._CURRENCY_FORMAT),
            ("Investimento em Capital de Giro", "working_capital_investment", self._CURRENCY_FORMAT),
        ]
        for row_idx, (label, field, fmt) in enumerate(wc_rows, start=2):
            values = [getattr(p, field) for p in periods]
            self._write_label_value_row(sheet2, row_idx, label, values, fmt)
        self._format_sheet("Cap 2 Capital de Giro")

        # ---------- Chapter 3: Outros Capitais (Other Capital) ----------
        sheet3 = self._wb.create_sheet("Cap 3 Outros Capitais")
        self._write_header_row(sheet3, row=1, values=header_values)

        oc_rows = [
            ("Imobilizado Líquido (PP&E)", "ppe_net", self._CURRENCY_FORMAT),
            ("Intangível Líquido", "intangibles_net", self._CURRENCY_FORMAT),
            ("Outros Capitais Líquido", "other_capital_net", self._CURRENCY_FORMAT),
            ("Fluxo de Caixa de Investimentos", "investing_cash_flow", self._CURRENCY_FORMAT),
        ]
        for row_idx, (label, field, fmt) in enumerate(oc_rows, start=2):
            values = [getattr(p, field) for p in periods]
            self._write_label_value_row(sheet3, row_idx, label, values, fmt)
        self._format_sheet("Cap 3 Outros Capitais")

        # ---------- Chapter 4: Financiamento (Funding + Cash Flow) ----------
        sheet4 = self._wb.create_sheet("Cap 4 Financiamento")
        self._write_header_row(sheet4, row=1, values=header_values)

        funding_rows = [
            ("Dívida Total", "total_debt", self._CURRENCY_FORMAT),
            ("Dívida Líquida", "net_debt", self._CURRENCY_FORMAT),
            ("Patrimônio Líquido", "shareholders_equity", self._CURRENCY_FORMAT),
            ("Dívida / PL", "debt_to_equity", '#,##0.00'),
            ("Fluxo de Caixa Operacional", "operating_cash_flow", self._CURRENCY_FORMAT),
            ("Fluxo de Caixa de Investimentos", "investing_cash_flow", self._CURRENCY_FORMAT),
            ("Fluxo de Caixa de Financiamento", "financing_cash_flow", self._CURRENCY_FORMAT),
            ("Fluxo de Caixa Líquido", "net_cash_flow", self._CURRENCY_FORMAT),
            ("Fluxo de Caixa Livre", "free_cash_flow", self._CURRENCY_FORMAT),
        ]
        for row_idx, (label, field, fmt) in enumerate(funding_rows, start=2):
            values = [getattr(p, field) for p in periods]
            self._write_label_value_row(sheet4, row_idx, label, values, fmt)
        self._format_sheet("Cap 4 Financiamento")

    def _add_power_of_one_sheet(self, analysis_result: AnalysisResult) -> None:
        """
        Add Power of One analysis sheet.

        Displays all levers with:
        - Lever name and category
        - Current value
        - Change amount and unit
        - Profit impact
        - Cash impact
        - Value impact
        """
        sheet = self._wb.create_sheet("Power of One")

        headers = [
            "Alavanca",
            "Categoria",
            "Valor Atual",
            "Variação",
            "Impacto Lucro",
            "Impacto Caixa",
            "Impacto Valor",
        ]
        self._write_header_row(sheet, row=1, values=headers)

        for row_idx, lever in enumerate(analysis_result.power_of_one, start=2):
            sheet.cell(row=row_idx, column=1, value=lever.label_pt)
            sheet.cell(row=row_idx, column=2, value=lever.category)

            current_val_cell = sheet.cell(row=row_idx, column=3, value=float(lever.current_value))
            current_val_cell.number_format = self._CURRENCY_FORMAT

            change_label = f"{float(lever.change_amount)} {lever.change_unit}"
            sheet.cell(row=row_idx, column=4, value=change_label)

            profit_cell = sheet.cell(row=row_idx, column=5, value=float(lever.profit_impact))
            profit_cell.number_format = self._CURRENCY_FORMAT

            cash_cell = sheet.cell(row=row_idx, column=6, value=float(lever.cash_impact))
            cash_cell.number_format = self._CURRENCY_FORMAT

            value_cell = sheet.cell(row=row_idx, column=7, value=float(lever.value_impact))
            value_cell.number_format = self._CURRENCY_FORMAT

        self._format_sheet("Power of One")

    def _add_financial_statements_sheets(self, analysis_result: AnalysisResult) -> None:
        """
        Add DRE, Fluxo de Caixa, Indicadores, and Qualidade do Caixa sheets.
        """
        periods = analysis_result.periods
        period_labels = [p.period for p in periods]
        header_values = ["Métrica"] + period_labels

        # ---------- DRE (Income Statement) ----------
        dre_sheet = self._wb.create_sheet("DRE")
        self._write_header_row(dre_sheet, row=1, values=header_values)

        dre_rows = [
            ("Receita Bruta", "gross_revenue", self._CURRENCY_FORMAT),
            ("(-) Devoluções e Deduções", "returns_deductions", self._CURRENCY_FORMAT),
            ("Receita Líquida", "net_revenue", self._CURRENCY_FORMAT),
            ("(-) CMV", "cogs", self._CURRENCY_FORMAT),
            ("Lucro Bruto", "gross_profit", self._CURRENCY_FORMAT),
            ("Margem Bruta %", "gross_margin_pct", self._PERCENT_FORMAT),
            ("(-) Despesas Operacionais", "operating_expenses", self._CURRENCY_FORMAT),
            ("EBITDA", "ebitda", self._CURRENCY_FORMAT),
            ("Margem EBITDA %", "ebitda_margin_pct", self._PERCENT_FORMAT),
            ("(-) Depreciação e Amortização", "depreciation_amortization", self._CURRENCY_FORMAT),
            ("EBIT", "ebit", self._CURRENCY_FORMAT),
            ("Margem EBIT %", "ebit_margin_pct", self._PERCENT_FORMAT),
            ("(-) Despesas Financeiras", "financial_expenses", self._CURRENCY_FORMAT),
            ("(+) Receitas Financeiras", "financial_income", self._CURRENCY_FORMAT),
            ("(+/-) Outros", "other_income_expenses", self._CURRENCY_FORMAT),
            ("EBT", "ebt", self._CURRENCY_FORMAT),
            ("(-) IRPJ", "irpj_tax", self._CURRENCY_FORMAT),
            ("(-) CSLL", "csll_tax", self._CURRENCY_FORMAT),
            ("Lucro Líquido", "net_income", self._CURRENCY_FORMAT),
            ("Margem Líquida %", "net_margin_pct", self._PERCENT_FORMAT),
        ]
        for row_idx, (label, field, fmt) in enumerate(dre_rows, start=2):
            values = [getattr(p, field) for p in periods]
            self._write_label_value_row(dre_sheet, row_idx, label, values, fmt)
        self._format_sheet("DRE")

        # ---------- Fluxo de Caixa (Cash Flow Statement) ----------
        cf_sheet = self._wb.create_sheet("Fluxo de Caixa")
        self._write_header_row(cf_sheet, row=1, values=header_values)

        cf_rows = [
            ("Fluxo de Caixa Operacional", "operating_cash_flow"),
            ("Fluxo de Caixa de Investimentos", "investing_cash_flow"),
            ("Fluxo de Caixa de Financiamento", "financing_cash_flow"),
            ("Fluxo de Caixa Líquido", "net_cash_flow"),
            ("Fluxo de Caixa Livre", "free_cash_flow"),
        ]
        for row_idx, (label, field) in enumerate(cf_rows, start=2):
            values = [getattr(p, field) for p in periods]
            self._write_label_value_row(cf_sheet, row_idx, label, values, self._CURRENCY_FORMAT)
        self._format_sheet("Fluxo de Caixa")

        # ---------- Indicadores (Ratios) ----------
        ind_sheet = self._wb.create_sheet("Indicadores")
        self._write_header_row(ind_sheet, row=1, values=header_values)

        ratio_rows = [
            ("Liquidez Corrente", "current_ratio", '#,##0.00'),
            ("Liquidez Seca", "quick_ratio", '#,##0.00'),
            ("ROE %", "roe_pct", self._PERCENT_FORMAT),
            ("ROA %", "roa_pct", self._PERCENT_FORMAT),
            ("ROCE %", "roce_pct", self._PERCENT_FORMAT),
            ("Dívida / PL", "debt_to_equity", '#,##0.00'),
        ]
        for row_idx, (label, field, fmt) in enumerate(ratio_rows, start=2):
            values = [getattr(p, field) for p in periods]
            self._write_label_value_row(ind_sheet, row_idx, label, values, fmt)
        self._format_sheet("Indicadores")

        # ---------- Qualidade do Caixa (Cash Quality Metrics) ----------
        cq_sheet = self._wb.create_sheet("Qualidade do Caixa")
        self._write_header_row(cq_sheet, row=1, values=["Métrica", "Valor", "Nota"])

        for row_idx, metric in enumerate(analysis_result.cash_quality, start=2):
            cq_sheet.cell(row=row_idx, column=1, value=metric.label_pt)
            val_cell = cq_sheet.cell(row=row_idx, column=2, value=float(metric.value))
            val_cell.number_format = self._CURRENCY_FORMAT
            cq_sheet.cell(row=row_idx, column=3, value=metric.grade)

        self._format_sheet("Qualidade do Caixa")
