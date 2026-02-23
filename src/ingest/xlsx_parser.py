"""Parser for Excel files containing ERP financial data."""

from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Optional

try:
    from openpyxl import load_workbook  # type: ignore[import-untyped]
except ImportError:  # pragma: no cover
    load_workbook = None  # handled at runtime

from src.models import AccountEntry


# Recognised header variants (lowered) → canonical name
_HEADER_ALIASES: Dict[str, str] = {
    "codigo": "code",
    "código": "code",
    "codigo da conta": "code",
    "código da conta": "code",
    "account code": "code",
    "descricao": "description",
    "descrição": "description",
    "description": "description",
    "saldo inicial": "opening_balance",
    "opening balance": "opening_balance",
    "debitos": "total_debits",
    "débitos": "total_debits",
    "debits": "total_debits",
    "creditos": "total_credits",
    "créditos": "total_credits",
    "credits": "total_credits",
    "saldo final": "closing_balance",
    "closing balance": "closing_balance",
    "periodo": "period",
    "período": "period",
    "period": "period",
}


class ERPExcelParser:
    """Parses Excel files containing ERP balancete and cash flow data."""

    def __init__(self, file_path: str, sheet_name: Optional[str] = None) -> None:
        self.file_path = Path(file_path)
        self.sheet_name = sheet_name

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _to_decimal(value) -> Decimal:
        if value is None:
            return Decimal("0")
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        text = str(value).strip()
        if not text:
            return Decimal("0")
        if "," in text and "." in text:
            text = text.replace(".", "").replace(",", ".")
        elif "," in text:
            text = text.replace(",", ".")
        return Decimal(text)

    @staticmethod
    def _detect_header_row(ws, max_scan: int = 20) -> tuple:
        """Return (row_index_0based, {canonical: col_index_0based})."""
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
            mapping: Dict[str, int] = {}
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                key = str(cell).strip().lower()
                canonical = _HEADER_ALIASES.get(key)
                if canonical:
                    mapping[canonical] = col_idx
            if "code" in mapping and ("closing_balance" in mapping or "opening_balance" in mapping):
                return row_idx, mapping
        raise ValueError("Could not detect header row in Excel sheet")

    def _get_worksheet(self):
        if load_workbook is None:
            raise ImportError("openpyxl is required for Excel parsing: pip install openpyxl")
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")
        wb = load_workbook(str(self.file_path), read_only=True, data_only=True)
        if self.sheet_name:
            ws = wb[self.sheet_name]
        else:
            ws = wb.active
        return wb, ws

    # ------------------------------------------------------------------
    # Balancete
    # ------------------------------------------------------------------

    def parse_balancete(self, skip_rows: int = 0) -> List[AccountEntry]:
        """Parse balancete sheet — auto-detects header row and columns."""
        wb, ws = self._get_worksheet()
        try:
            header_row, col_map = self._detect_header_row(ws)
            entries: List[AccountEntry] = []
            data_start = header_row + 2 + skip_rows  # 1-based for openpyxl

            for row in ws.iter_rows(min_row=data_start, values_only=True):
                code_val = row[col_map["code"]] if "code" in col_map else None
                if code_val is None or str(code_val).strip() == "":
                    continue

                entries.append(
                    AccountEntry(
                        code=str(code_val).strip(),
                        description=str(row[col_map["description"]]).strip() if "description" in col_map and row[col_map["description"]] else "",
                        opening_balance=self._to_decimal(row[col_map["opening_balance"]]) if "opening_balance" in col_map else Decimal("0"),
                        total_debits=self._to_decimal(row[col_map["total_debits"]]) if "total_debits" in col_map else Decimal("0"),
                        total_credits=self._to_decimal(row[col_map["total_credits"]]) if "total_credits" in col_map else Decimal("0"),
                        closing_balance=self._to_decimal(row[col_map["closing_balance"]]) if "closing_balance" in col_map else Decimal("0"),
                        period=str(row[col_map["period"]]).strip() if "period" in col_map and row[col_map["period"]] else "",
                    )
                )
            return entries
        finally:
            wb.close()

    # ------------------------------------------------------------------
    # Fluxo de caixa
    # ------------------------------------------------------------------

    def parse_fluxo_caixa(self, skip_rows: int = 0) -> List[AccountEntry]:
        """Parse fluxo de caixa sheet from Excel."""
        wb, ws = self._get_worksheet()
        try:
            header_row, col_map = self._detect_header_row(ws)
            entries: List[AccountEntry] = []
            data_start = header_row + 2 + skip_rows

            for row in ws.iter_rows(min_row=data_start, values_only=True):
                code_val = row[col_map.get("code", 0)]
                if code_val is None or str(code_val).strip() == "":
                    continue

                valor = self._to_decimal(row[col_map["closing_balance"]]) if "closing_balance" in col_map else Decimal("0")
                entries.append(
                    AccountEntry(
                        code=str(code_val).strip(),
                        description=str(row[col_map["description"]]).strip() if "description" in col_map and row[col_map["description"]] else "",
                        opening_balance=Decimal("0"),
                        total_debits=Decimal("0"),
                        total_credits=Decimal("0"),
                        closing_balance=valor,
                        period=str(row[col_map["period"]]).strip() if "period" in col_map and row[col_map["period"]] else "",
                    )
                )
            return entries
        finally:
            wb.close()
